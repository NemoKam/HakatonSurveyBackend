import uuid
import random
import string
import hashlib
from datetime import timedelta, datetime, timezone

import pandas as pd
import jwt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fastapi import Depends, Header, Response
from sqlalchemy.ext.asyncio import AsyncSession

import config
from fastapp import crud, schemas, exceptions, models
from fastapp.database import get_db


def generate_random_string(length: int = random.randint(1, 128), only_digits: bool = False) -> str:
    chars = string.digits

    if not only_digits:
        chars += string.ascii_uppercase

    return ''.join(random.choice(chars) for _ in range(length))


def hash_password(password: str) -> str:
    password_bytes = password.encode('utf-8')
    hash_object = hashlib.sha256(password_bytes)

    return hash_object.hexdigest()


def _get_utc_now() -> datetime:
    current_utc_time = datetime.now(timezone.utc)

    return current_utc_time


def _create_token(payload: dict, expire: datetime) -> schemas.JwtTokenCreate:
    payload[config.EXP] = expire

    token = schemas.JwtTokenCreate(
        token=jwt.encode(payload, config.JWT_SECRET,
                         algorithm=config.JWT_ALGORITHM),
        payload=payload,
        expire=expire,
    )

    return token


def _create_access_token(payload: dict, minutes: int | None = None) -> schemas.JwtTokenCreate:
    expire = _get_utc_now() + timedelta(
        minutes=minutes or config.ACCESS_TOKEN_EXPIRES_MINUTES
    )
    
    access_token: schemas.JwtTokenCreate = _create_token(payload, expire)

    return access_token


def _create_refresh_token(payload: dict) -> schemas.JwtTokenCreate:
    expire = _get_utc_now() + timedelta(
        minutes=config.REFRESH_TOKEN_EXPIRES_MINUTES
    )

    refresh_token: schemas.JwtTokenCreate = _create_token(payload, expire)

    return refresh_token


def create_token_pair(user_id: uuid.UUID) -> schemas.TokenPair:
    payload = {config.SUB: str(user_id), config.JTI: str(
        uuid.uuid4()), config.IAT: _get_utc_now()}

    return schemas.TokenPair(
        access=_create_access_token(payload={**payload}),
        refresh=_create_refresh_token(payload={**payload}),
    )


def decode_access_token(token: str) -> dict:
    payload = jwt.decode(token, config.JWT_SECRET,
                         algorithms=[config.JWT_ALGORITHM])

    return payload


def refresh_token_state(token: str) -> schemas.JwtTokenGet:
    payload = jwt.decode(token, config.JWT_SECRET,
                         algorithms=[config.JWT_ALGORITHM])

    access_token = _create_access_token(payload=payload)

    return access_token


def add_refresh_token_cookie(response: Response, token: str):
    exp = _get_utc_now() + timedelta(minutes=config.REFRESH_TOKEN_EXPIRES_MINUTES)
    exp.replace(tzinfo=timezone.utc)

    response.set_cookie(
        key="refresh_token",
        value=token,
        expires=int(exp.timestamp()),
        httponly=True,
    )


def remove_refresh_token_from_cookie(response: Response):
    response.delete_cookie(
        key="refresh_token"
    )


def create_excel_with_data(file_name: str, data_main: list[dict], data_tasks: list[dict]):
    df_main = pd.DataFrame(data_main)

    with pd.ExcelWriter(f"{file_name}.xlsx", engine="openpyxl") as writer:
        df_main.to_excel(writer, sheet_name="Основная таблица", index=False)

        for task_num, task_entries in enumerate(data_tasks, start=1):
            task_data = []

            for entry in task_entries:
                row = [entry["email"]] + entry["answers"]
                task_data.append(row)

            max_answers = max(len(entry["answers"]) for entry in task_entries)
            columns = ["email"] + [f"Ответ {i + 1}" for i in range(max_answers)]

            df_task = pd.DataFrame(task_data, columns=columns)
            sheet_name = f"Задание #{task_num}"
            df_task.to_excel(writer, sheet_name=sheet_name, index=False)
    
    wb = load_workbook(f"{file_name}.xlsx")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[row[0].row].height = 30  # Устанавливаем высоту строки
        
        for col in range(1, ws.max_column + 1):
            max_length = max(len(str(ws.cell(row=row, column=col).value)) for row in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = max_length + 2  # Добавляем небольшой отступ


    wb.save(f"{file_name}.xlsx")


async def refresh_survey_document(db: AsyncSession, survey_document_id: uuid.UUID, survey_document_title: str):
    survey_document: models.SurveyDocument = await crud.get_survey_document_by_id(db, survey_document_id)
    survey_id: uuid.UUID = survey_document.survey_id

    survey: models.Survey | None = await crud.get_survey_by_id(db, survey_id, True)

    if survey:
        file_name: str = f"{config.SURVEY_DOCUMENT_SAVE_PATH}{survey_document_title}"
        document_main_data: list[dict] = []
        document_data_tasks: list[dict] = []
        questions_info: dict = {}

        questions: list[models.Question] = survey.questions

        for question in questions:
            questions_info[question.id] = {
                "correct_answers": list(sorted([question_answer.text for question_answer in question.answers if question_answer.is_correct])),
                "type": question.type,
                "score": question.score if question.score else 0,
                "user_answers_info": []
            }
        
        for user_survey_result in survey.user_survey_results:
            user: models.User | None = user_survey_result.user
            user_score: int = 0

            user_info = {
                "email": user.email if user else "",
                "name": user.name if user else "",
                "surname": user.surname if user else "",
            }

            for user_question_result in user_survey_result.user_questions:
                curent_question: models.Question = user_question_result.question
                question_info: dict = questions_info[curent_question.id]

                user_answers: list[str] = list(sorted([user_answer_result.text for user_answer_result in user_question_result.user_answers]))
                
                if question_info["type"] in ["text", "choose_one", "dropdown_list"]:
                    if len(user_answers) == 1 and user_answers[0] in question_info["correct_answers"]:
                        user_score += question_info["score"]
                elif question_info["type"] == "choose_many":
                    if user_answers == question_info["correct_answers"]:
                        user_score += question_info["score"]
                
                user_answers_info = {
                    "email": user_info["email"],
                    "answers": user_answers
                }

                question_info["user_answers_info"].append(user_answers_info)

            user_info["score"] = user_score

            document_main_data.append(user_info)
            
        for question_id in questions_info:
            users_answers = questions_info[question_id]["user_answers_info"]
            document_data_tasks.append(users_answers)
        
        create_excel_with_data(file_name, document_main_data, document_data_tasks)


async def get_user_from_access_token(authorization: str | None = Header(None), db: AsyncSession = Depends(get_db)) -> models.User:
    if not authorization:
        raise exceptions.AuthFailedException(detail="Authorization header missing")

    if not authorization.startswith("Bearer "):
        raise exceptions.AuthFailedException(detail="Invalid authorization header format")

    access_token = authorization.split(" ")[1]

    try:
        user_info = decode_access_token(access_token)
    except jwt.exceptions.DecodeError:
        raise exceptions.AuthFailedException(detail="Invalid access_token")
    except jwt.exceptions.ExpiredSignatureError:
        raise exceptions.AuthFailedException(detail="Access token expired")

    user_id: uuid.UUID = user_info[config.SUB]

    user = await crud.get_user_by_id(db, user_id)

    if not user:
        raise exceptions.AuthFailedException(detail="Invalid access_token")

    return user


async def check_survey_is_valid(db: AsyncSession, survey: models.Survey, authorization: str | None = Header(None)):
    if survey and not survey.is_anonim:
        if survey.expire_datetime and survey.expire_datetime < datetime.now():
            raise exceptions.NotAllowedException(detail="Survey is finished")
        user: models.User = await get_user_from_access_token(authorization, db)
        
        if not survey.send_multiple_times:
            user_passed_survey_ids: list[uuid.UUID] = [user_survey_result.user_id for user_survey_result in survey.user_survey_results]
            if user.id in user_passed_survey_ids:
                raise exceptions.NotAllowedException(detail="You are already passed this survey")

